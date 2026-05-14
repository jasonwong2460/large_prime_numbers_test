import gmpy2
import streamlit as st
import pandas as pd
import numpy as np
import time
import psutil
import sys
import os
import matplotlib.pyplot as plt
from datetime import datetime
from generate_test_numbers import generate_test_numbers
from scipy.interpolate import PchipInterpolator

# 添加算法模块路径
sys.path.append(os.path.dirname(__file__))
from algorithms import *

# 页面配置
st.set_page_config(
    page_title="素性测试算法对比评测",
    page_icon="🔢",
    layout="wide"
)

# 初始化会话状态
if 'test_results' not in st.session_state:
    st.session_state.test_results = []
if 'test_numbers' not in st.session_state:
    st.session_state.test_numbers = None


def run_algorithm_test(algorithm_name, algorithm_func, numbers, test_type):
    """运行单个算法的测试"""
    results = []
    times = []
    memories = []

    progress_bar = st.progress(0)
    status_text = st.empty()

    for i, num in enumerate(numbers):
        # 提取数字和具体类型/位数
        if isinstance(num, tuple):
            num_val, meta = num
            if isinstance(meta, int):
                # 针对素数和合数，meta 是位数
                actual_test_type = f"{test_type}-{meta}位"
            else:
                # 针对特殊数，meta 是特殊数名称字符串
                actual_test_type = meta
        else:
            num_val = num
            actual_test_type = test_type

        status_text.markdown(f"""
        **正在测试:** `{algorithm_name}`  
        **进度:** `{i + 1}/{len(numbers)}`  
        **数字:** `{str(num_val)}`
        """)

        process = psutil.Process()
        mem_before = process.memory_info().rss / 1024 / 1024
        start_time = time.time()

        result = algorithm_func(num_val)

        # 转换为毫秒 (ms)
        elapsed = (time.time() - start_time) * 1000
        mem_after = process.memory_info().rss / 1024 / 1024
        mem_used = max(0, mem_after - mem_before)

        results.append({
            'number': num_val,
            'is_prime': result,
            'time': elapsed,  # 现在存储的是毫秒
            'memory': mem_used,
            'test_type': actual_test_type
        })

        times.append(elapsed)
        memories.append(mem_used)

        progress_bar.progress((i + 1) / len(numbers))

    status_text.text(f"{algorithm_name} 测试完成")
    return results, times, memories


def compare_algorithms():
    """对比所有算法"""
    st.header("🔬 算法性能对比")

    if st.session_state.test_numbers is None:
        st.warning("请先在侧边栏生成测试数据")
        return

    test_numbers = st.session_state.test_numbers

    # 算法列表
    algorithms = {
        "试除法": trial_division_test,
        "费马测试": fermat_test,
        "米勒-拉宾": miller_rabin_test,
        "Solovay_Strassen方法": solovay_strassen_test,
        "卢卡斯-莱默(梅森数专用)": lucas_lehmer_test,
        "Baillie_PSW方法": baillie_psw_test,
        "AKS算法": aks_test,
        "APR-CL算法": APRtest,
        "试除法(优化版)": trial_division_optimized_test,
        "费马测试(优化版)": fermat_optimized_test,
        "卢卡斯-莱默(优化版)": lucas_lehmer_optimized_test
    }

    selected_algorithms = st.multiselect(
        "选择要对比的算法",
        list(algorithms.keys()),
        default=list(algorithms.keys())[:3]
    )

    if st.button("开始性能对比测试"):
        all_results = {}
        comparison_data = []
        special_details = {}

        for algo_name in selected_algorithms:
            st.subheader(f"测试 {algo_name}")

            tab1, tab2, tab3 = st.tabs(["素数测试", "合数测试", "特殊数测试"])

            with tab1:
                st.write("测试素数")
                results, _, _ = run_algorithm_test(
                    algo_name, algorithms[algo_name], test_numbers['primes'], "素数"
                )
                all_results[f"{algo_name}_primes"] = results

                # 按具体的位数组分别统计
                df_res = pd.DataFrame(results)
                if not df_res.empty:
                    for t_type, group in df_res.groupby('test_type'):
                        correct = 0
                        for _, r in group.iterrows():
                            if r['is_prime'] == gmpy2.is_prime(r['number']):
                                correct += 1
                        accuracy = (correct / len(group)) * 100 if len(group) > 0 else 0
                        comparison_data.append({
                            '算法': algo_name,
                            '类型': t_type,
                            '平均时间(毫秒)': group['time'].mean(),
                            '最大时间(毫秒)': group['time'].max(),
                            '平均内存(MB)': group['memory'].mean(),
                            '准确率(%)': accuracy
                        })

            with tab2:
                st.write("测试合数")
                results, _, _ = run_algorithm_test(
                    algo_name, algorithms[algo_name], test_numbers['composites'], "合数"
                )
                all_results[f"{algo_name}_composites"] = results

                df_res = pd.DataFrame(results)
                if not df_res.empty:
                    for t_type, group in df_res.groupby('test_type'):
                        correct = 0
                        for _, r in group.iterrows():
                            if r['is_prime'] == gmpy2.is_prime(r['number']):
                                correct += 1
                        accuracy = (correct / len(group)) * 100 if len(group) > 0 else 0
                        comparison_data.append({
                            '算法': algo_name,
                            '类型': t_type,
                            '平均时间(毫秒)': group['time'].mean(),
                            '最大时间(毫秒)': group['time'].max(),
                            '平均内存(MB)': group['memory'].mean(),
                            '准确率(%)': accuracy
                        })

            with tab3:
                st.write("测试特殊数")
                results, times, memories = run_algorithm_test(
                    algo_name, algorithms[algo_name], test_numbers['special'], "特殊数"
                )
                all_results[f"{algo_name}_special"] = results

                correct = sum(1 for r in results if r['is_prime'] == gmpy2.is_prime(r['number']))
                accuracy = (correct / len(results)) * 100 if results else 0
                comparison_data.append({
                    '算法': algo_name,
                    '类型': '特殊数',
                    '平均时间(毫秒)': np.mean(times) if times else 0,
                    '最大时间(毫秒)': np.max(times) if times else 0,
                    '平均内存(MB)': np.mean(memories) if memories else 0,
                    '准确率(%)': accuracy
                })

                # 记录特殊数详细数据表
                carmichael_correct, carmichael_total = 0, 0
                pseudoprime_correct, pseudoprime_total = 0, 0
                mersenne_correct, mersenne_total = 0, 0
                carmichael_times, pseudoprime_times, mersenne_times = [], [], []

                for r in results:
                    standard = gmpy2.is_prime(r['number'])
                    if r['test_type'] == "卡迈克尔数":
                        carmichael_total += 1
                        carmichael_times.append(r['time'])
                        if r['is_prime'] == standard: carmichael_correct += 1
                    elif r['test_type'] == "强伪素数":
                        pseudoprime_total += 1
                        pseudoprime_times.append(r['time'])
                        if r['is_prime'] == standard: pseudoprime_correct += 1
                    elif r['test_type'] == "梅森数":
                        mersenne_total += 1
                        mersenne_times.append(r['time'])
                        if r['is_prime'] == standard: mersenne_correct += 1

                special_details[algo_name] = {
                    '卡迈克尔数': {
                        '正确数': carmichael_correct, '总数': carmichael_total,
                        '准确率(%)': (carmichael_correct / carmichael_total * 100) if carmichael_total > 0 else 0,
                        '平均时间(毫秒)': np.mean(carmichael_times) if carmichael_times else 0
                    },
                    '强伪素数': {
                        '正确数': pseudoprime_correct, '总数': pseudoprime_total,
                        '准确率(%)': (pseudoprime_correct / pseudoprime_total * 100) if pseudoprime_total > 0 else 0,
                        '平均时间(毫秒)': np.mean(pseudoprime_times) if pseudoprime_times else 0
                    },
                    '梅森数': {
                        '正确数': mersenne_correct, '总数': mersenne_total,
                        '准确率(%)': (mersenne_correct / mersenne_total * 100) if mersenne_total > 0 else 0,
                        '平均时间(毫秒)': np.mean(mersenne_times) if mersenne_times else 0
                    }
                }

        # 显示对比结果
        display_comparison_results(comparison_data)

        st.header("📊 特殊数详细对比")
        if special_details:
            summary_data = []
            for algo_name, details in special_details.items():
                for special_type, stats in details.items():
                    if stats['总数'] > 0:
                        summary_data.append({
                            '算法': algo_name,
                            '特殊数类型': special_type,
                            '正确数/总数': f"{stats['正确数']}/{stats['总数']}",
                            '准确率(%)': f"{stats['准确率(%)']:.2f}",
                            '平均时间(毫秒)': f"{stats['平均时间(毫秒)']:.4f}"
                        })
            if summary_data:
                st.dataframe(pd.DataFrame(summary_data), hide_index=True, use_container_width=True)

        # ========== 下载详细测试结果 ==========
        st.header("📥 下载详细测试结果")

        all_details = []
        for key, results in all_results.items():
            parts = key.split('_')
            algo_name = parts[0]
            for r in results:
                standard_result = gmpy2.is_prime(r['number'])
                is_correct = (r['is_prime'] == standard_result)
                all_details.append({
                    '算法': algo_name,
                    '测试类型': r['test_type'],
                    '测试数字': str(r['number']),
                    '算法结果': '素数' if r['is_prime'] else '合数',
                    '标准答案': '素数' if standard_result else '合数',
                    '是否正确': '✓ 正确' if is_correct else '✗ 错误',
                    '执行时间(毫秒)': r['time'],
                    '内存占用(MB)': r['memory']
                })

        if all_details:
            df_details = pd.DataFrame(all_details)
            import io
            from openpyxl.styles import PatternFill

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_details.to_excel(writer, index=False, sheet_name='详细测试结果')
                worksheet = writer.sheets['详细测试结果']
                red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

                correct_col = None
                for col_idx, col_name in enumerate(df_details.columns, start=1):
                    if col_name == '是否正确':
                        correct_col = col_idx
                        break

                if correct_col:
                    for row_idx in range(2, len(df_details) + 2):
                        if worksheet.cell(row=row_idx, column=correct_col).value == '✗ 错误':
                            for col_idx in range(1, len(df_details.columns) + 1):
                                worksheet.cell(row=row_idx, column=col_idx).fill = red_fill

            excel_data = output.getvalue()
            st.download_button(
                label="下载详细测试结果 (Excel)",
                data=excel_data,
                file_name=f"detailed_test_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


from matplotlib.ticker import LogFormatter

class MyLogFormatter(LogFormatter):
    def fix_minus(self, s):
        # 强制将 Unicode 的数学减号替换为标准的 ASCII 减号
        return s.replace('\u2212', '-')


def display_comparison_results(comparison_data):
    """显示对比结果"""
    import re
    st.header("📈 性能对比图表")

    # 中文字体兼容性配置
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans', 'sans-serif']
    plt.rcParams['axes.unicode_minus'] = False

    df = pd.DataFrame(comparison_data)

    # 核心排序逻辑：解析类型字符串，用于X轴严格排序
    def get_sort_key(val):
        val = str(val)
        # 确定大类优先级: 0=素数, 1=合数, 2=特殊数
        if "素数" in val and "伪" not in val and "梅森" not in val:
            category = 0
        elif "合数" in val:
            category = 1
        else:
            category = 2  # 特殊数及其他

        # 提取位数 (如果有)
        match = re.search(r'(\d+)位', val)
        bits = int(match.group(1)) if match else 0

        # 返回排序元组 (类别升序, 位数升序, 名称兜底)
        return (category, bits, val)


    # 数据清洗：从“类型”字符串中提取数字位宽 (例如 "素数-256位" -> 256)
    def extract_bits(s):
        match = re.search(r'(\d+)位', str(s))
        return int(match.group(1)) if match else None

    df['bits'] = df['类型'].apply(extract_bits)

    # 过滤出线性增长数据，并剔除 bits 为空或数据异常的行
    df_linear = df[df['bits'].notnull()].copy()

    if df_linear.empty:
        st.warning("⚠️ 当前测试数据不包含位宽信息，无法生成趋势图。请先在侧边栏生成数据。")
        return

    # 绘图配置矩阵
    plot_configs = [
        ('平均时间(毫秒)', '算法执行耗时趋势', '时间 (ms)', True),
        ('平均内存(MB)', '算法内存占用趋势', '内存 (MB)', False),
        ('准确率(%)', '算法准确率稳定性', '准确率 (%)', False)
    ]

    for col, title, ylabel, use_log in plot_configs:
        st.subheader(f"● {title}")
        fig, ax = plt.subplots(figsize=(12, 6))

        # 遍历分类（实线表示素数，虚线表示合数）
        for test_category in ['素数', '合数']:
            category_df = df_linear[df_linear['类型'].str.contains(test_category)].copy()

            if category_df.empty:
                continue

            for algo in category_df['算法'].unique():
                algo_df = category_df[category_df['算法'] == algo]

                # --- 关键：数据聚合与清洗 ---
                # 必须先求均值，再按位宽排序，确保 X 轴是唯一的递增序列
                clean_df = algo_df.groupby('bits')[col].mean().reset_index().dropna().sort_values('bits')

                x = clean_df['bits'].values
                y = clean_df[col].values

                # 排除 y 轴无效数据
                if use_log:
                    mask = y >= 0
                    x, y = x[mask], y[mask]

                if len(x) < 2:
                    continue

                # 执行插值拟合
                if len(x) >= 4:
                    try:
                        x_new = np.linspace(x.min(), x.max(), 300)
                        # 保形插值，不会超出数据范围，无需 clip
                        pchip = PchipInterpolator(x, y)
                        y_smooth = pchip(x_new)

                        if use_log:
                            # PCHIP 保证非负，但若原始数据最小值为正，这里不会产生 ≤0 的值
                            # 安全起见可加保护
                            y_smooth = np.maximum(y_smooth, 1e-12)

                        line_style = '-' if test_category == '素数' else '--'
                        ax.plot(x_new, y_smooth, label=f"{algo} ({test_category})",
                                linestyle=line_style, linewidth=2, alpha=0.8)
                    except Exception:
                        # 保形插值极少失败，失败时可回退到简单折线
                        ax.plot(x, y, 'o-', alpha=0.6, label=f"{algo} ({test_category})")
                else:
                    # 点数 < 4 直接连线
                    ax.plot(x, y, 'o-', alpha=0.6, label=f"{algo} ({test_category})")

        # 图表格式化
        if use_log:
            ax.set_yscale('log')
            # 强制使用自定义格式化器，解决负号方块问题
            ax.yaxis.set_major_formatter(MyLogFormatter())

        ax.set_xlabel('二进制位数 (Bits)', fontweight='bold')
        ax.set_ylabel(ylabel, fontweight='bold')
        ax.grid(True, which="both", ls=":", alpha=0.5)

        if '准确率' in col:
            ax.set_ylim(-5, 105)

        if '内存' in col:
            ax.set_ylim(0, 0.001)

        # 调整图例位置
        ax.legend(bbox_to_anchor=(1.01, 1), loc='upper left', fontsize='small')
        plt.tight_layout()
        st.pyplot(fig)

    # 数据明细展示
    with st.expander("查看原始统计数据表格"):
        st.dataframe(df.drop(columns=['bits']), use_container_width=True, hide_index=True)

    st.subheader("详细测试数据")
    # 为 DataFrame 添加一个隐藏的排序列以便正确排序表格
    df['sort_key'] = df['类型'].apply(get_sort_key)
    # 先按算法排序，再按我们自定义的顺序严格排列
    df_sorted = df.sort_values(by=['算法', 'sort_key']).drop(columns=['sort_key'])
    st.dataframe(df_sorted, use_container_width=True, hide_index=True)


def single_number_test():
    """单个数测试"""
    st.header("🔢 单个数字测试")

    col1, col2 = st.columns([2, 1])

    with col1:
        test_number = st.text_input(
            "输入要测试的数字",
            placeholder="例如: 123456789"
        )

    with col2:
        algorithms_selected = st.multiselect(
            "选择要使用的算法",
            ["试除法", "费马测试", "米勒-拉宾", "Baillie_PSW方法", "Solovay_Strassen方法",
             "卢卡斯-莱默(梅森数专用)", "AKS算法", "APR-CL算法", "试除法(优化版)", "费马测试(优化版)", "卢卡斯-莱默(优化版)"],
            default=["米勒-拉宾", "Baillie_PSW方法"]
        )

    if st.button("开始测试") and test_number:
        try:
            num = int(test_number)
            st.write(f"### 测试数字: {num}")

            algorithms_map = {
                "试除法": trial_division_test,
                "费马测试": fermat_test,
                "米勒-拉宾": miller_rabin_test,
                "Solovay_Strassen方法": solovay_strassen_test,
                "卢卡斯-莱默(梅森数专用)": lucas_lehmer_test,
                "Baillie_PSW方法": baillie_psw_test,
                "AKS算法": aks_test,
                "APR-CL算法": APRtest,
                "试除法(优化版)": trial_division_optimized_test,
                "费马测试(优化版)": fermat_optimized_test,
                "卢卡斯-莱默(优化版)": lucas_lehmer_optimized_test
            }

            results_data = []

            for algo_name in algorithms_selected:
                if algo_name in algorithms_map:
                    with st.spinner(f"运行 {algo_name}..."):
                        start_time = time.time()
                        result = algorithms_map[algo_name](num)

                        # 转换为毫秒 (ms)
                        elapsed = (time.time() - start_time) * 1000

                        results_data.append({
                            "算法": algo_name,
                            "结果": "素数" if result else "合数",
                            "执行时间(毫秒)": f"{elapsed:.4f}"
                        })

            df = pd.DataFrame(results_data)
            st.dataframe(df,
                         hide_index=True,
                         use_container_width=True,
                         column_config={
                             "算法": st.column_config.TextColumn(width="small"),
                             "结果": st.column_config.TextColumn(width="small"),
                             "执行时间(毫秒)": st.column_config.TextColumn(width="medium")
                         })

        except ValueError:
            st.error("请输入有效的整数")


def main():
    """主函数"""
    st.title("🔢 素性测试算法对比评测系统")
    st.markdown("---")

    st.sidebar.title("⚙️ 配置面板")

    generate_test_numbers()

    tab1, tab2 = st.tabs(["📊 算法对比测试", "🔬 单个数测试"])

    with tab1:
        compare_algorithms()

    with tab2:
        single_number_test()


if __name__ == "__main__":
    main()